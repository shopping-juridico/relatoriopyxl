from xls2xlsx import XLS2XLSX
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from datetime import datetime
import os

def convert():
    x2x = XLS2XLSX("excel files/grade-exportacao.xls")
    x2x.to_xlsx("excel files/convertido-suporte.xlsx")

def copia_lista(col):
    wb = load_workbook("excel files/convertido-suporte.xlsx")
    ws = wb.active
    
    lista_col = []
    for cell in ws['{}'.format(col)]:
        celula = str(cell.value)
        lista_col.append(celula)
    return lista_col

wb2 = Workbook()
ws2 = wb2.active

def cola_lista(lista, coluna):
    r = 1
    for cell in lista:
        ws2.cell(row=r, column=coluna).value = cell
        r += 1

def formata_suporte():
    convert()
    id = copia_lista('C')
    acao_tipo = copia_lista('D')
    duracao = copia_lista('K')
    ultimo_andamento = copia_lista('L')
    agente = copia_lista('O')
    descricao = copia_lista('P')

    cola_lista(id, 1)
    cola_lista(acao_tipo, 2)
    cola_lista(agente, 3)
    cola_lista(duracao, 4)
    cola_lista(ultimo_andamento, 5)
    cola_lista(descricao, 6)

    ws2.title = "Suportes TR"

    # excluir linhas sem "TR - " na coluna J
    for cell in ws2['F']:
        if(cell.value is not None):
            if cell.value == "Descrição":
                continue 
            if 'TR - ' not in cell.value:
                ws2.delete_rows(cell.row)
        else:
            ws2.delete_rows(cell.row) 

    # formatando como tabela e definindo um estilo
    tab = Table(displayName="Table1", ref="A1:F{}".format(ws2.max_row - 1)) #CRM inclui 1 linha a mais

    style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws2.add_table(tab)

    # fonte
    arial_font = Font(name="Arial")
    i = 1
    while i <= ws2.max_row:
        for cell in ws2[i]:
            cell.font = arial_font
        i = i + 1

    # ocultar grid
    ws2.sheet_view.showGridLines
    True
    ws2.sheet_view.showGridLines = False

    # auto ajuste de largura das células
    for column_cells in ws2.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws2.column_dimensions[column_cells[0].column_letter].width = length * 1.3

    # para ajustar altura das células
    #(alguma coisa assim):
    #for row_cells in ws.rows:
    #ws.row_dimensions[0].height = 70

    # ocultar linhas e colunas
    # ws.column_dimensions.group('A','D', hidden=True)
    # ws.row_dimensions.group(1, 5, hidden=True)

    ws2.delete_rows(ws2.max_row)

    now = datetime.now()

    wb2.save('excel files/Suportes TR {}.{}.xlsx'.format(now.day, now.month))

    os.remove("excel files/convertido-suporte.xlsx")