from openpyxl import Workbook
wb = Workbook() #cria a variável da classe

ws1 = wb.create_sheet("Mysheet") #cria a planilha e define um título para ela
ws2 = wb.create_sheet("Mysheet")
ws3 = wb.create_sheet("My Third Sheet")

ws2.title = "My Second Sheet" # altera o título da planilha

ws3.sheet_properties.tabColor = "1072BA" # altera a cor da aba da planilha

wb.save("/home/vitor/projetos/int-py-xl/mytest.xlsx")

# próximo passo é obter o nome de uma sheet como uma chave da workbook

ws3 = wb["My Third Sheet"]

print(wb.sheetnames)