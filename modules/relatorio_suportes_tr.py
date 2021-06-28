from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from datetime import datetime

def formata_suporte():

    # carregar o arquivo fonte
    wb = load_workbook('excel files/grade-exportacao.xlsx')

    # criar a variável da planilha que será manipulada
    ws = wb.worksheets[0]

    # nome da planilha
    ws.title = "Relatório"

    # excluindo colunas
    ws.delete_cols(2) 
    ws.delete_cols(6, 2)
    ws.delete_cols(7)
    ws.delete_cols(9, 2)
    ws.delete_cols(11, 3)

    # excluir linhas sem "TR - " na coluna J
    for cell in ws['J']:
        if(cell.value is not None):
            if cell.value == "Descrição":
                continue 
            if 'TR - ' not in cell.value:
                ws.delete_rows(cell.row)
        else:
            ws.delete_rows(cell.row) 

    # formatando como tabela e definindo um estilo
    tab = Table(displayName="Table1", ref="A1:J{}".format(ws.max_row - 1)) #CRM inclui 1 linha a mais

    style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # fonte
    arial_font = Font(name="Arial")
    i = 1
    while i <= ws.max_row:
        for cell in ws[i]:
            cell.font = arial_font
        i = i + 1
        
    # ocultar grid
    ws.sheet_view.showGridLines
    True
    ws.sheet_view.showGridLines = False

    # auto ajuste de largura das células
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length * 1.3

    # para ajustar altura das células
    #(alguma coisa assim):
    #for row_cells in ws.rows:
    #ws.row_dimensions[0].height = 70

    # ocultar linhas e colunas
    # ws.column_dimensions.group('A','D', hidden=True)
    # ws.row_dimensions.group(1, 5, hidden=True)

    #data atual
    now = datetime.now()

    wb.save("excel files/Relatório Suportes TR - {}.0{}.xlsx".format(now.day, now.month))
