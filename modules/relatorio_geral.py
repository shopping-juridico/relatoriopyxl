from xls2xlsx import XLS2XLSX
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

def convert():
    x2x = XLS2XLSX("excel files/grade-exportacao.xls")
    x2x.to_xlsx("excel files/convertido-geral.xlsx")

def copia_lista(col):
    wb = load_workbook("excel files/convertido-geral.xlsx")
    ws = wb.active
    
    lista_col = []
    for cell in ws['{}'.format(col)]:
        celula = str(cell.value)
        lista_col.append(celula)
    return lista_col

wb2 = Workbook()
ws2 = wb2.active

def cola_lista(lista, coluna):
    r = 1
    for cell in lista:
        ws2.cell(row=r, column=coluna).value = cell
        r += 1

def formata_geral():
    convert()
    id = copia_lista('C')
    tipo = copia_lista('D')
    data_inicio = copia_lista('I')
    ultimo_andamento = copia_lista('L')
    agente = copia_lista('O')
    descricao = copia_lista('P')

    cola_lista(id, 1)
    cola_lista(tipo, 2)
    cola_lista(agente, 3)
    cola_lista(data_inicio, 4)
    cola_lista(ultimo_andamento, 5)
    cola_lista(descricao, 6)

    ws2.delete_rows(ws2.max_row)

    now = datetime.now()

    wb2.save('excel files/Suportes Geral {}.{}.xlsx'.format(now.day, now.month))

    os.remove("excel files/convertido-geral.xlsx")