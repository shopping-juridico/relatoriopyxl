from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from datetime import datetime

wb = load_workbook('grade-exportacao (1).xlsx')

ws = wb.worksheets[0]

ws.title = "Relatório Andamentos"
print(wb.sheetnames)
print(ws.max_row, "linhas e", ws.max_column, "colunas")

for cell in ws['G']:
    if(cell.value is not None):
        if cell.value == "Data Cadastro":
            continue
        if cell.value < ws.cell(row=2, column=7).value:
            ws.delete_rows(cell.row)
    else:
        ws.delete_rows(cell.row)

tab = Table(displayName="Table1", ref="A1:K{}".format(ws.max_row - 1))
style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)

arial_font = Font(name="Arial")
i = 1
while i <= ws.max_row:
    for cell in ws[i]:
        cell.font = arial_font
    i = i + 1

ws.sheet_view.showGridLines
True
ws.sheet_view.showGridLines = False

now = datetime.now()

wb.save("/home/vitor/projetos/relatorios_py-xl/Relatório de Andamentos - Vitor - {}.0{}.xlsx".format(now.day, now.month))
