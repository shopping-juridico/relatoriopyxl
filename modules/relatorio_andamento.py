from xls2xlsx import XLS2XLSX
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from datetime import datetime

def convert():
    x2x = XLS2XLSX("excel files/grade-exportacao.xls")
    x2x.to_xlsx("excel files/convertido-andamentos.xlsx")

def copia_lista(col):
    wb = load_workbook("excel files/convertido-andamentos.xlsx")
    ws = wb.active
    
    lista_col = []
    for cell in ws['{}'.format(col)]:
        celula = str(cell.value)
        lista_col.append(celula)
    return lista_col

wb2 = Workbook()
ws2 = wb2.active

def cola_lista(lista, coluna):
    r = 1
    for cell in lista:
        ws2.cell(row=r, column=coluna).value = cell
        r += 1

def formata_andamentos():
    convert()
    concluida = copia_lista('A')
    acao = copia_lista('B')
    acao_tipo = copia_lista('C')
    responsavel = copia_lista('D')
    agente = copia_lista('E')
    contato = copia_lista('F')
    data_cadastro = copia_lista('G')
    data = copia_lista('H')
    tempo_min = copia_lista('I')
    descricao = copia_lista('K')

    cola_lista(concluida, 1)
    cola_lista(acao, 2)
    cola_lista(acao_tipo, 3)
    cola_lista(responsavel, 4)
    cola_lista(agente, 5)
    cola_lista(contato, 6)
    cola_lista(data_cadastro, 7)
    cola_lista(data, 8)
    cola_lista(tempo_min, 9)
    cola_lista(descricao, 10)

    ws2.delete_rows(ws2.max_row)

    for cell in ws2['G']:
        if(cell.value is not None):
            if cell.value == "Data Cadastro":
                continue
            if cell.value < ws2.cell(row=2, column=7).value:
                ws2.delete_rows(cell.row)
        else:
            ws2.delete_rows(cell.row)

    tab = Table(displayName="Table1", ref="A1:J{}".format(ws2.max_row - 1))
    style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws2.add_table(tab)

    arial_font = Font(name="Arial")
    i = 1
    while i <= ws2.max_row:
        for cell in ws2[i]:
            cell.font = arial_font
        i = i + 1      

    ws2.sheet_view.showGridLines
    True
    ws2.sheet_view.showGridLines = False

    for column_cells in ws2.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws2.column_dimensions[column_cells[0].column_letter].width = length * 1.3

    now = datetime.now()

    wb2.save('excel files/generated/Relatório de Andamentos - Vitor - {}.{}.xlsx'.format(now.day, now.month))

#formata_andamentos()